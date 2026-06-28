"""Exportación del estado de resultados (P&L) a Excel y PDF."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Iterable

from app.schemas.reports_financial import PnlAccountRow, ProfitAndLossResponse

_PNL_SECTIONS: tuple[tuple[str, str, str | None], ...] = (
    ("ingresos", "Ingresos", "Total ingresos"),
    ("otros_ingresos", "Otros ingresos", None),
    ("costo_de_ventas", "Costo de Ventas", "Total costo de ventas"),
    ("gastos", "Gastos", "Total gastos"),
    ("otros_gastos_financieros", "Otros gastos financieros", None),
)


def _rows_for_section(report: ProfitAndLossResponse, attr: str) -> list[PnlAccountRow]:
    return list(getattr(report, attr, None) or [])


def _sum_rows(rows: Iterable[PnlAccountRow]) -> Decimal:
    total = Decimal("0")
    for row in rows:
        total += Decimal(str(row.monto))
    return total


def _flatten_account_rows(rows: list[PnlAccountRow], depth: int = 0) -> list[tuple[str, Decimal]]:
    out: list[tuple[str, Decimal]] = []
    indent = "  " * depth
    for row in rows:
        out.append((f"{indent}{row.cuenta}", Decimal(str(row.monto))))
        subs = row.subcuentas or []
        if subs:
            out.extend(_flatten_account_rows(subs, depth + 1))
    return out


def _build_export_rows(report: ProfitAndLossResponse) -> list[tuple[str, Decimal | None]]:
    """Filas planas: (concepto, monto). ``None`` en monto = fila de encabezado de sección."""
    lines: list[tuple[str, Decimal | None]] = []

    for attr, title, total_label in _PNL_SECTIONS:
        section_rows = _rows_for_section(report, attr)
        if not section_rows:
            continue
        lines.append((title.upper(), None))
        lines.extend(_flatten_account_rows(section_rows))
        if total_label:
            lines.append((total_label, _sum_rows(section_rows)))
        lines.append(("", None))

    lines.append(("UTILIDAD BRUTA (Ingresos − Costo de ventas)", report.utilidad_bruta))
    lines.append(("", None))
    lines.append(("UTILIDAD NETA", report.utilidad_neta))
    return lines


def pnl_export_filename(report: ProfitAndLossResponse, ext: str) -> str:
    start = report.start_date.isoformat()
    end = report.end_date.isoformat()
    if start[:7] == end[:7]:
        period = start[:7]
    else:
        period = f"{start}_{end}"
    return f"Perdidas_Ganancias_{period}.{ext}"


def build_pnl_excel_bytes(report: ProfitAndLossResponse) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("Para exportar Excel instala openpyxl en el servidor.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Pérdidas y ganancias"

    ws["A1"] = "Estado de resultados (P&L)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Periodo: {report.start_date.isoformat()} — {report.end_date.isoformat()}"
    ws["A3"] = "Consolidado en USD"
    ws.merge_cells("A1:B1")
    ws.merge_cells("A2:B2")
    ws.merge_cells("A3:B3")

    header_row = 5
    ws.cell(row=header_row, column=1, value="Concepto").font = Font(bold=True)
    ws.cell(row=header_row, column=2, value="Monto").font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for col in (1, 2):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = header_row + 1
    for concept, amount in _build_export_rows(report):
        if concept == "" and amount is None:
            row_idx += 1
            continue
        concept_cell = ws.cell(row=row_idx, column=1, value=concept)
        if amount is None:
            concept_cell.font = Font(bold=True)
            row_idx += 1
            continue
        ws.cell(row=row_idx, column=2, value=float(amount)).number_format = '"$"#,##0.00'
        if concept.startswith(("Total ", "UTILIDAD")):
            concept_cell.font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
        row_idx += 1

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pnl_pdf_bytes(report: ProfitAndLossResponse) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError("Para exportar PDF instala reportlab en el servidor.") from exc

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=0.65 * inch,
        rightMargin=0.65 * inch,
        topMargin=0.65 * inch,
        bottomMargin=0.65 * inch,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Pérdidas y ganancias", styles["Title"]),
        Paragraph(
            f"Periodo: {report.start_date.isoformat()} — {report.end_date.isoformat()}",
            styles["Normal"],
        ),
        Paragraph("Consolidado en USD", styles["Normal"]),
        Spacer(1, 0.2 * inch),
    ]

    table_data: list[list[str]] = [["Concepto", "Monto"]]
    bold_row_indexes: set[int] = {0}

    for concept, amount in _build_export_rows(report):
        if concept == "" and amount is None:
            continue
        if amount is None:
            table_data.append([concept, ""])
            bold_row_indexes.add(len(table_data) - 1)
            continue
        formatted = f"${float(amount):,.2f}"
        table_data.append([concept, formatted])
        if concept.startswith(("Total ", "UTILIDAD")):
            bold_row_indexes.add(len(table_data) - 1)

    table = Table(table_data, colWidths=[4.4 * inch, 1.5 * inch], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ]
        )
    )
    for idx in bold_row_indexes:
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"),
                ]
            )
        )

    story.append(table)
    doc.build(story)
    return buf.getvalue()
